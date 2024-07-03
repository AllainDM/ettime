import openpyxl
from datetime import timedelta


all_list = [["https://us.gblnet.net/oper/?core_section=task&action=show&id="],
            ["Номер заявки", "Время назначенное КО", "Время подключения", "Причина переноса"]]

counter_all = 0
counter_before = 0
counter_self = 0
counter_after = 0


def read():

    workbook_north = openpyxl.load_workbook('north.xlsx')
    workbook_south = openpyxl.load_workbook('south.xlsx')
    workbook_kolpino = openpyxl.load_workbook('kolpino.xlsx')
    workbook_west = openpyxl.load_workbook('west.xlsx')
    workbook_east = openpyxl.load_workbook('east.xlsx')
    sheet_north = workbook_north.active
    sheet_south = workbook_south.active
    sheet_kolpino = workbook_kolpino.active
    sheet_west = workbook_west.active
    sheet_east = workbook_east.active

    # 47 ЛС
    # 48 Назначенная дата
    # 55 Дата выполнения

    for row in sheet_north.iter_rows(values_only=True):
        t_o = "Север"
        read_one(row, t_o)

    for row in sheet_south.iter_rows(values_only=True):
        t_o = "Юг"
        read_one(row, t_o)

    for row in sheet_kolpino.iter_rows(values_only=True):
        t_o = "Юг"
        read_one(row, t_o)

    for row in sheet_west.iter_rows(values_only=True):
        t_o = "Запад"
        read_one(row, t_o)

    for row in sheet_east.iter_rows(values_only=True):
        t_o = "Восток"
        read_one(row, t_o)


def read_one(row, t_o):
    global counter_all
    global counter_self
    global counter_after
    global counter_before

    print("##################################################################")
    comment = ''
    try:
        start_time = row[48]
        print(f"start_time {start_time}")
        print(type(start_time))
        start_time_str = str(start_time)
        print(type(start_time))
        # print(f"start_time[8:10] {start_time[8:10]}")
        start_time_d = start_time_str[8:10]
        print(f"start_time_d = start_time[8:10] {start_time_d}")
        start_time_m = start_time_str[-5:-3]
        print(f"start_time_m = start_time[-5:-3] {start_time_m}")

        end_time = row[55]
        print(f"end_time {end_time}")
        end_time_str = str(end_time)
        end_time_d = end_time_str[8:10]
        print(f"end_time_d= end_time[8:10] {end_time_d}")
        end_time_m = end_time_str[-5:-3]
        print(f"end_time_m = end_time[-5:-3] {end_time_m}")

        str_address = str(row[49])  # Получаем строку из NoneType
        lst_address = list(str_address.split(","))  # Делаем из строки список
        new_lst_address = lst_address[2:]  # Убираем первые два элемента
        new_str_address = ''.join(new_lst_address)  # Обратно создаем строку из списка

        print(f"Проверяем время")
        if start_time_str[-5:-3] == "00":
            # if start_time_m == end_time_m and start_time_d > end_time_d:
            #     print('start_time_m == end_time_m and start_time_d > end_time_d')
            #     comment = "Подключили раньше"
            #     print(comment)
            #     counter_all += 1
            #     counter_before += 1
            #     all_list.append([t_o, row[47], row[48], row[55], comment])
            #     # all_list.append([t_o, row[47], row[48], row[55], comment, new_str_address])
            if start_time - timedelta(hours=2) > end_time:
                print("start_time_m > end_time_m")
                comment = "Подключили раньше"
                print(comment)
                counter_all += 1
                counter_before += 1
                all_list.append([t_o, row[47], row[48], row[55], comment])
                # all_list.append([t_o, row[47], row[48], row[55], comment, new_str_address])
            # elif start_time_m == end_time_m and start_time_d == end_time_d:
            #     counter_all += 1
            #     comment = "Все норм?"
            #     print(comment)
            #     # all_list.append([t_o, row[47], row[48], row[55], comment])
            #     # all_list.append([t_o, row[47], row[48], row[55], comment, new_str_address])
            elif start_time + timedelta(hours=4) > end_time:
                comment = "Подключили в более удобное время для клиента"
                print(comment)
                counter_all += 1
                counter_after += 1
                all_list.append([t_o, row[47], row[48], row[55], comment])
                # all_list.append([t_o, row[47], row[48], row[55], comment, new_str_address])
        else:
            # comment = " "
            comment = "Назначили сами"
            print(comment)
            counter_all += 1
            counter_self += 1

            all_list.append([t_o, row[47], row[48], row[55], comment])
            # all_list.append([t_o, row[47], row[48], row[55], comment, new_str_address])

        # if type(start_time) is str:
        #     if start_time[-2:] == "00":
        #         comment = "Была назначена"
    except TypeError:
        print("Ошибка")
    print(f"ЛС {row[47]}")
    # print(f"Адрес {new_str_address}")
    print(f"Назначенная дата {row[48]}")
    print(f"Дата выполнения {row[55]}")
    print(f"Комментарий {comment}")
    # link = f"https://us.gblnet.net/oper/index.php?core_section=task&action=show&id={row[47]}"
    # link1 = f'=ГИПЕРССЫЛКА({link};{row[47]})'
    # link2 = f"=ГИПЕРССЫЛКА(CONCAT($A$1;{row[47]});{row[47]})"
    # print(link2)
    # ws.write(1, 24, "https://us.gblnet.net/oper/?core_section=task&action=show&id=")
    # all_list.append([t_o, row[47], row[48], row[55], comment])


read()


def write():
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append(["Всего: ", counter_all])
    sheet.append(["Сделано раньше: ", counter_before])
    sheet.append(["Назначили сами: ", counter_self])
    sheet.append(["Перенесено: ", counter_after])

    for row in all_list:
        sheet.append(row)

    workbook.save('new_file.xlsx')


write()
